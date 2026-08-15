"""日別シート形式のスケジュールExcelから、配送/現場タスク候補の中間CSVを作る。

自由記述のExcel (例: 2026年8月スケジュール.xlsx) を完全自動で解釈するのは無理なので、
「人が見て補正できる中間CSV」を作ることが目的です。取りこぼしや誤判定は
出力CSVを手で直す前提で、判定できなかった行は notes に残します。

使い方:
    pip install -r requirements.txt
    python3 src/extract_tasks_from_excel.py \
        --input "2026年8月スケジュール.xlsx" \
        --output data/tasks_2026-08.csv

    # 特定日だけ
    python3 src/extract_tasks_from_excel.py --input ... --sheet 815 --sheet 820

前提にしているシート構造:
    A1                  その日の日付
    2行目               各列の担当者/チャンネル名 (例: 「タンCh(ｱﾄﾞｸﾘｴｲﾂ様)」)
    A列                 案件記入欄。11行1ブロックのテンプレート
                        (お客様名＠場所 / 営業名 / 内容 / 入り時間 / OP人数 /
                         設営撤去+◯名 / 車両 / 備考)
    C列以降             担当者ごとの当日の動き。「＠」を含むセルが1タスクの先頭で、
                        以降の行に営業名・内容・時刻・人数・車両・備考が続く
    「車両」列          連絡メモ用なのでタスクとしては読まない
"""

import argparse
import csv
import datetime as dt
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT_FILE = BASE_DIR / "data" / "extracted_tasks.csv"

OUTPUT_COLUMNS = [
    "date",
    "task_type",
    "customer",
    "venue_name",
    "address",
    "start_time",
    "end_time",
    "required_vehicle",
    "required_staff_count",
    "assigned_vehicle",
    "assigned_staff",
    "origin",
    "notes",
]

# A列の案件記入欄は11行で1ブロック。先頭ブロックは3行目から。
SALES_BLOCK_FIRST_ROW = 3
SALES_BLOCK_HEIGHT = 11
SALES_BLOCK_FIELDS = {
    0: "customer_venue",
    1: "sales_person",
    2: "content",
    3: "arrival",
    5: "op_staff",
    6: "setup_staff",
    8: "vehicle",
    9: "notes",
}

# 未入力のままテンプレートの見出しが残っているセルは空扱いにする。
TEMPLATE_LABELS = {
    "お客様名＠場所",
    "営業名",
    "内容(例：映像/音響/ﾓﾆﾀｰ/LED等)",
    "入り時間",
    "OP人数(例：映像OP◯名/音響OP◯名)",
    "設営撤去+◯名",
    "車両",
    "備考",
    "案件記入欄（ｲﾍﾞﾝﾄ部営業のみ）",
}

# 【】内の記号をタスク種別にする。【設/OP/撤】のような複合は
# 「設営/オペレート/撤去」のようにそのまま並べて人が見て判断できるようにする。
TASK_TYPE_BY_TAG_PART = {
    "設": "設営",
    "撤": "撤去",
    "OP": "オペレート",
    "RH": "リハーサル",
    "立": "立会",
    "搬入": "搬入",
    "搬出": "搬出",
    "納": "納品",
    "回": "回収",
    "引": "引取",
    "引取": "引取",
    "返": "返却",
    "修理出し": "修理出し",
    "修理引取": "修理引取",
    "調整": "調整",
    "前日移動": "前日移動",
    "仮": "仮押さえ",
}

TAG_PATTERN = re.compile(r"【([^】]*)】")
TIME_PATTERN = re.compile(r"^(\d{1,2})[:：時](\d{2})?")
VEHICLE_PATTERN = re.compile(
    r"(\d+(?:\.\d+)?t(?:\s*\+\s*\d+(?:\.\d+)?t)*|1BOX|1ﾎﾞｯｸｽ|ﾊｲﾙｰﾌ|ハイルーフ|ﾜﾝﾎﾞｯｸｽ)",
    re.IGNORECASE,
)
ASSIGNED_VEHICLE_PATTERN = re.compile(r"(\d+号車)")
STAFF_COUNT_PATTERN = re.compile(r"(?:設営撤去|設営|撤去|立会|立ち合い)[^0-9]{0,4}(\d+)\s*名")
MOVEMENT_PATTERN = re.compile(r"(ｾﾝﾀｰ発|センター発|ｾﾝﾀｰ出社|センター出社|ｾﾝﾀｰ移動|センター移動|直行|直帰)")
SHEET_NAME_PATTERN = re.compile(r"^(\d{1,2})(\d{1,2})$")
# 「10:00 ～ 19:00」のような作業可能帯と、「(11:30)」のような指定時刻。
TIME_RANGE_PATTERN = re.compile(r"(\d{1,2}:\d{2})\s*[～~〜-]\s*(\d{1,2}:\d{2})")
APPOINTMENT_PATTERN = re.compile(r"[(（](\d{1,2}:\d{2})[)）]")
# 「A12345678(◯◯会場)」のような伝票番号＋場所。
ORDER_CODE_VENUE_PATTERN = re.compile(r"^([A-Za-z]?\d{4,})[(（](.+)[)）]$")


def normalize(value):
    """セル値を1行の文字列にする。空セルは空文字。"""
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.strftime("%H:%M") if value.year == 1900 else value.strftime("%Y-%m-%d")
    if isinstance(value, dt.time):
        return value.strftime("%H:%M")
    text = str(value).replace("\u3000", " ")
    text = " / ".join(line.strip() for line in text.splitlines() if line.strip())
    return text.strip()


def is_meaningful(text):
    return bool(text) and text not in TEMPLATE_LABELS


def parse_time(text):
    """「10:50」「7時入り」「20：00」などから HH:MM を取り出す。"""
    normalized = unicodedata.normalize("NFKC", text)
    match = TIME_PATTERN.match(normalized)
    if not match:
        return ""

    hour = int(match.group(1))
    minute = int(match.group(2) or 0)
    if hour > 23 or minute > 59:
        return ""
    return f"{hour:02d}:{minute:02d}"


def sheet_date(worksheet, sheet_name, year, month):
    """A1の日付を使い、無ければシート名 (例: 815 = 8月15日) から日付を作る。"""
    value = worksheet["A1"].value
    if isinstance(value, dt.datetime):
        return value.date().isoformat()

    match = SHEET_NAME_PATTERN.match(sheet_name)
    if match and year:
        sheet_month, day = int(match.group(1)), int(match.group(2))
        try:
            return dt.date(year, month or sheet_month, day).isoformat()
        except ValueError:
            return ""
    return ""


def split_customer_venue(text):
    """「【設】◯◯様＠△△会場南1-2」→ (◯◯様, △△会場南1-2)"""
    body = TAG_PATTERN.sub("", text).strip()
    body = re.sub(r"^\[[^\]]*\]", "", body).strip()

    for separator in ("＠", "@"):
        if separator in body:
            customer, venue = body.split(separator, 1)
            return customer.strip(), venue.strip()
    return body, ""


def task_type_from_tag(text):
    parts = []
    for tag in TAG_PATTERN.findall(text):
        for part in tag.split("/"):
            part = re.sub(r"\d+$", "", part.strip())
            parts.append(TASK_TYPE_BY_TAG_PART.get(part, part))

    parts = [part for part in dict.fromkeys(parts) if part]
    return "/".join(parts) if parts else "要確認"


def column_staff(worksheet, column):
    return normalize(worksheet.cell(row=2, column=column).value)


def is_vehicle_column(worksheet, column):
    """「車両」列は連絡メモ用なのでタスクとして読まない。"""
    return any(
        normalize(worksheet.cell(row=row, column=column).value) == "車両"
        for row in (1, 2)
    )


def last_data_row(worksheet):
    for row in range(worksheet.max_row, 0, -1):
        for column in range(1, worksheet.max_column + 1):
            if normalize(worksheet.cell(row=row, column=column).value):
                return row
    return 0


def has_customer_marker(text):
    return "＠" in text or "@" in text


def is_placeholder(customer, venue):
    """【納】＠A1234() のような凡例行をタスクとして拾わない。"""
    return not customer and (not venue or "()" in venue)


def collect_details(worksheet, column, start_row, end_row):
    """タスク見出しの次の行から、次の見出し (または空行の連続) までを集める。"""
    details = []
    empty_streak = 0

    for row in range(start_row + 1, end_row + 1):
        text = normalize(worksheet.cell(row=row, column=column).value)
        if not text:
            empty_streak += 1
            if empty_streak >= 3:
                break
            continue

        if has_customer_marker(text):
            break

        empty_streak = 0
        details.append(text)

    return details


def nearest_movement(worksheet, column, anchor_row):
    """見出しの上下で一番近い「ｾﾝﾀｰ発 / 直行 / 直帰」などの行を返す。"""
    for offset in range(1, 6):
        for row in (anchor_row - offset, anchor_row + offset):
            if row < 2:
                continue
            text = normalize(worksheet.cell(row=row, column=column).value)
            if text and MOVEMENT_PATTERN.search(text):
                return text
    return ""


def build_task(date, task_type, customer, venue, details, assigned_staff, movement):
    """判定できた項目を列に振り分け、残りは notes に入れる。"""
    times = []
    vehicles = []
    staff_counts = []
    notes = []

    appointments = []
    end_times = []

    for text in details:
        normalized = unicodedata.normalize("NFKC", text)
        time_value = parse_time(text)
        if time_value and len(normalized) <= 8:
            times.append(time_value)
            continue

        range_match = TIME_RANGE_PATTERN.search(normalized)
        if range_match:
            times.append(range_match.group(1))
            end_times.append(range_match.group(2))

        appointment_match = APPOINTMENT_PATTERN.search(normalized)
        if appointment_match:
            appointments.append(appointment_match.group(1))

        vehicle_match = VEHICLE_PATTERN.search(normalized)
        if vehicle_match:
            vehicles.append(vehicle_match.group(1))

        staff_match = STAFF_COUNT_PATTERN.search(normalized)
        if staff_match:
            staff_counts.append(int(staff_match.group(1)))

        notes.append(text)

    movement_text = movement or ""
    assigned_vehicle = ""
    assigned_vehicle_match = ASSIGNED_VEHICLE_PATTERN.search(movement_text)
    if assigned_vehicle_match:
        assigned_vehicle = assigned_vehicle_match.group(1)

    if "直行" in movement_text and "ｾﾝﾀｰ" not in movement_text and "センター" not in movement_text:
        origin = "直行"
    elif MOVEMENT_PATTERN.search(movement_text):
        origin = "センター"
    else:
        origin = ""

    if appointments:
        notes.append(f"指定時刻: {appointments[0]}")

    if movement_text:
        notes.append(f"移動: {movement_text}")

    order_code_match = ORDER_CODE_VENUE_PATTERN.match(venue)
    if order_code_match:
        notes.append(f"伝票: {order_code_match.group(1)}")
        venue = order_code_match.group(2)

    # 指定時刻があればそれを到着時刻とし、無ければ作業可能帯の開始を使う。
    start_time = appointments[0] if appointments else (times[0] if times else "")
    if end_times:
        end_time = end_times[0]
    else:
        end_time = times[1] if len(times) > 1 else ""

    return {
        "date": date,
        "task_type": task_type,
        "customer": customer,
        "venue_name": venue,
        "address": "",
        "start_time": start_time,
        "end_time": end_time,
        "required_vehicle": " ".join(dict.fromkeys(vehicles)),
        "required_staff_count": str(max(staff_counts)) if staff_counts else "",
        "assigned_vehicle": assigned_vehicle,
        "assigned_staff": assigned_staff,
        "origin": origin,
        "notes": " / ".join(notes),
    }


def extract_sales_entries(worksheet, date, end_row):
    """A列の案件記入欄 (11行1ブロック) を読む。"""
    tasks = []

    for top in range(SALES_BLOCK_FIRST_ROW, end_row + 1, SALES_BLOCK_HEIGHT):
        values = {}
        for offset, field in SALES_BLOCK_FIELDS.items():
            text = normalize(worksheet.cell(row=top + offset, column=1).value)
            if is_meaningful(text):
                values[field] = text

        header = values.get("customer_venue", "")
        if not header or not has_customer_marker(header):
            continue

        customer, venue = split_customer_venue(header)
        if is_placeholder(customer, venue):
            continue

        details = [
            values[field]
            for field in ("content", "arrival", "op_staff", "setup_staff", "vehicle", "notes")
            if field in values
        ]
        task = build_task(date, "案件", customer, venue, details, "", "")
        if "sales_person" in values:
            task["notes"] = " / ".join(
                filter(None, [f"営業: {values['sales_person']}", task["notes"]])
            )
        tasks.append(task)

    return tasks


def extract_column_tasks(worksheet, date, column, end_row):
    """担当者列のタスク (「＠」を含む見出し行) を読む。"""
    tasks = []
    assigned_staff = column_staff(worksheet, column)

    for row in range(3, end_row + 1):
        header = normalize(worksheet.cell(row=row, column=column).value)
        if not header or not has_customer_marker(header):
            continue

        customer, venue = split_customer_venue(header)
        if is_placeholder(customer, venue):
            continue

        details = collect_details(worksheet, column, row, end_row)
        movement = nearest_movement(worksheet, column, row)
        tasks.append(
            build_task(
                date,
                task_type_from_tag(header),
                customer,
                venue,
                details,
                assigned_staff,
                movement,
            )
        )

    return tasks


def extract_sheet(worksheet, sheet_name, year, month):
    date = sheet_date(worksheet, sheet_name, year, month)
    end_row = last_data_row(worksheet)
    if end_row == 0:
        return []

    tasks = extract_sales_entries(worksheet, date, end_row)
    for column in range(3, worksheet.max_column + 1):
        if is_vehicle_column(worksheet, column):
            continue
        tasks.extend(extract_column_tasks(worksheet, date, column, end_row))

    return tasks


def extract_workbook(input_file, sheets=None, year=None, month=None):
    workbook = load_workbook(input_file, data_only=True, read_only=False)
    targets = sheets or workbook.sheetnames

    tasks = []
    for sheet_name in targets:
        if sheet_name not in workbook.sheetnames:
            raise SystemExit(f"シートが見つかりません: {sheet_name}")
        tasks.extend(extract_sheet(workbook[sheet_name], sheet_name, year, month))

    return tasks


def write_csv(tasks, output_file):
    output_file.parent.mkdir(parents=True, exist_ok=True)
    with output_file.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        writer.writerows(tasks)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True, help="スケジュールExcel (.xlsx)")
    parser.add_argument(
        "--output",
        default=str(DEFAULT_OUTPUT_FILE),
        help=f"出力CSV (既定: {DEFAULT_OUTPUT_FILE})",
    )
    parser.add_argument(
        "--sheet",
        action="append",
        dest="sheets",
        help="対象シート名。複数指定可 (既定: 全シート)",
    )
    parser.add_argument("--year", type=int, help="A1に日付が無いシート用の年")
    parser.add_argument("--month", type=int, help="A1に日付が無いシート用の月")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="CSVを書かずに件数と先頭数件だけ表示する",
    )
    args = parser.parse_args()

    tasks = extract_workbook(Path(args.input), args.sheets, args.year, args.month)

    if args.dry_run:
        for task in tasks[:5]:
            print(task)
        print(f"{len(tasks)}件を抽出しました (dry-run)。")
        return

    write_csv(tasks, Path(args.output))
    print(f"{len(tasks)}件を {args.output} に書き出しました。")
    print("内容を人が確認・補正してから、住所付与やルート最適化に進んでください。")


if __name__ == "__main__":
    main()
